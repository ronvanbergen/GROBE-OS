from __future__ import annotations

import os
import re
from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import Any, Optional

from fastapi import Depends, FastAPI, File, HTTPException, UploadFile
from fastapi.responses import HTMLResponse
from pydantic import BaseModel
from sqlalchemy import Column, DateTime, Float, Integer, String, Text, create_engine, or_
from sqlalchemy.orm import Session, declarative_base, sessionmaker
from sqlalchemy.sql import func
from openpyxl import load_workbook

APP_VERSION = "0.8.0-live-start"
DATA_DIR = Path(os.getenv("GROBE_OS_DATA_DIR", "/tmp/grobe_os_data"))
DATA_DIR.mkdir(parents=True, exist_ok=True)
DATABASE_URL = os.getenv("DATABASE_URL", f"sqlite:///{DATA_DIR / 'grobe_os.db'}")

engine = create_engine(DATABASE_URL, connect_args={"check_same_thread": False} if DATABASE_URL.startswith("sqlite") else {})
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
Base = declarative_base()


class Item(Base):
    __tablename__ = "items"
    id = Column(Integer, primary_key=True, index=True)
    code = Column(String, unique=True, index=True, nullable=False)
    omschrijving = Column(String, nullable=False)
    type = Column(String, nullable=False)
    eenheid = Column(String, default="stuks")
    voorraad = Column(Float, default=0)
    minimum = Column(Float, default=0)
    kostprijs = Column(Float, default=0)
    leverancier = Column(String, default="")
    gewijzigd = Column(DateTime, server_default=func.now(), onupdate=func.now())


class Movement(Base):
    __tablename__ = "movements"
    id = Column(Integer, primary_key=True, index=True)
    item_code = Column(String, index=True, nullable=False)
    mutatie = Column(Float, nullable=False)
    bron = Column(String, nullable=False)
    referentie = Column(String)
    gebruiker = Column(String, default="Ron")
    datum = Column(DateTime, server_default=func.now())


class ImportLog(Base):
    __tablename__ = "import_logs"
    id = Column(Integer, primary_key=True, index=True)
    source = Column(String, nullable=False)
    filename = Column(String, nullable=False)
    rows_read = Column(Integer, default=0)
    rows_changed = Column(Integer, default=0)
    message = Column(Text, default="")
    datum = Column(DateTime, server_default=func.now())


class Customer(Base):
    __tablename__ = "customers"
    id = Column(Integer, primary_key=True, index=True)
    naam = Column(String, unique=True, nullable=False)
    email = Column(String, default="")
    afleveradres = Column(Text, default="")


class Supplier(Base):
    __tablename__ = "suppliers"
    id = Column(Integer, primary_key=True, index=True)
    naam = Column(String, unique=True, nullable=False)
    type = Column(String, default="")
    email = Column(String, default="")


class PriceHistory(Base):
    __tablename__ = "price_history"
    id = Column(Integer, primary_key=True, index=True)
    item_code = Column(String, index=True, nullable=False)
    supplier = Column(String, default="")
    old_price = Column(Float, default=0)
    new_price = Column(Float, default=0)
    source = Column(String, default="")
    reference = Column(String, default="")
    datum = Column(DateTime, server_default=func.now())


class PurchaseLine(Base):
    __tablename__ = "purchase_lines"
    id = Column(Integer, primary_key=True, index=True)
    item_code = Column(String, index=True, nullable=False)
    description = Column(String, default="")
    quantity = Column(Float, default=0)
    unit_price = Column(Float, default=0)
    supplier = Column(String, default="")
    filename = Column(String, default="")
    raw_row = Column(Text, default="")
    datum = Column(DateTime, server_default=func.now())


def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


KLEUREN = {
    "ruim_voldoende": "#2E3C7A",
    "voldoende": "#8FBCE7",
    "bestellen": "#F5A623",
    "bijna_op": "#D32F2F",
}

TYPE_GROEP = {
    "GEREED_PRODUCT": "finished_stock",
    "TANK": "tanks",
    "CAPACITEIT": "production_capacity",
    "GRONDSTOF": "raw_materials",
    "VERPAKKING": "packaging",
    "DOOS": "packaging",
    "ETIKET": "labels",
}

ALIASES = {
    "BDG": ["bdg", "butyldiglycol", "butyl diglycol"],
    "AMPHOLAK": ["ampholak", "yjh"],
    "KOH": ["koh", "kaliumhydroxide", "potassium hydroxide"],
    "DISSOLVINE": ["dissolvine", "gl47", "gl-47"],
    "NL8P4": ["nl8p4", "rokanol nl8p4"],
    "GA8W": ["ga8w", "rokanol ga8w"],
    "NCS": ["ncs", "cumene", "natriumcumeensulfonaat", "hoesch ncs"],
    "DOOS-5L-UNI": ["doos universol 4 x 5", "doos 4 x 5", "4x5", "4 x 5 liter"],
    "DOOS-1L-UNI": ["doos universol 12 x 1", "doos 12 x 1", "12x1", "12 x 1 liter"],
    "FLES-600": ["600 ml witte fles", "fles 600", "600ml fles"],
    "TRIGGER-LB": ["lichtblauwe trigger", "trigger lichtblauw", "blue trigger"],
    "ETI-UNI-600": ["etiket universol 600", "label universol 600"],
    "UNI-5L": ["universol 5 liter", "universol 5l"],
    "UNI-1L": ["universol 1 liter", "universol 1l"],
    "UNI-600": ["universol 600 ml", "universol spray"],
}


def voorraad_status(item: Item) -> str:
    if item.minimum <= 0:
        return "ruim_voldoende"
    if item.voorraad <= item.minimum:
        return "bijna_op"
    if item.voorraad <= item.minimum * 2:
        return "bestellen"
    if item.voorraad <= item.minimum * 4:
        return "voldoende"
    return "ruim_voldoende"


def item_dict(r: Item) -> dict:
    st = voorraad_status(r)
    return {
        "id": r.id,
        "code": r.code,
        "omschrijving": r.omschrijving,
        "type": r.type,
        "eenheid": r.eenheid,
        "voorraad": r.voorraad,
        "minimum": r.minimum,
        "kostprijs": r.kostprijs or 0,
        "leverancier": r.leverancier or "",
        "status": st,
        "kleur": KLEUREN[st],
    }


def upsert_item(db: Session, **data):
    row = db.query(Item).filter(Item.code == data["code"]).first()
    if row:
        for k, v in data.items():
            setattr(row, k, v)
    else:
        db.add(Item(**data))


def seed():
    Base.metadata.create_all(bind=engine)
    db = SessionLocal()
    try:
        if db.query(Item).count() == 0:
            seed_items = [
                ("UNI-5L", "Universol 5 liter", "GEREED_PRODUCT", "stuks", 1552, 800, 0, "Distrifill"),
                ("UNI-1L", "Universol 1 liter", "GEREED_PRODUCT", "stuks", 8640, 2500, 0, "Distrifill"),
                ("UNI-600", "Universol 600 ml spray", "GEREED_PRODUCT", "stuks", 1440, 720, 0, "Distrifill"),
                ("UNI-DOEK", "Universol doekjes bus", "GEREED_PRODUCT", "stuks", 360, 120, 0, "Distrifill"),
                ("UNI-TANK", "Universol concentraat in tank", "TANK", "liter", 6300, 2000, 0, "Distrifill"),
                ("UNI-CAP-MICHAEL", "Productiecapaciteit Universol volgens Michael", "CAPACITEIT", "liter", 37800, 10000, 0, "Michael"),
                ("BDG", "Butyldiglycol", "GRONDSTOF", "kg/l", 0, 500, 0, "Julius Hoesch"),
                ("AMPHOLAK", "Ampholak", "GRONDSTOF", "kg/l", 0, 500, 0, "Nouryon"),
                ("KOH", "KOH", "GRONDSTOF", "kg/l", 0, 250, 0, "Julius Hoesch"),
                ("DISSOLVINE", "Dissolvine", "GRONDSTOF", "kg/l", 0, 250, 0, "Nouryon"),
                ("NL8P4", "ROKAnol NL8P4", "GRONDSTOF", "kg/l", 0, 250, 0, "PCC"),
                ("GA8W", "ROKAnol GA8W", "GRONDSTOF", "kg/l", 0, 250, 0, "PCC"),
                ("NCS", "Hoesch NCS", "GRONDSTOF", "kg/l", 0, 250, 0, "Julius Hoesch"),
                ("DOOS-5L-UNI", "Doos Universol 4 x 5 liter", "DOOS", "stuks", 1320, 300, 0, "Superdoos"),
                ("DOOS-1L-UNI", "Doos Universol 12 x 1 liter", "DOOS", "stuks", 1320, 300, 0, "Superdoos"),
                ("FLES-600", "600 ml witte fles", "VERPAKKING", "stuks", 2240, 500, 0, "Distrifill"),
                ("TRIGGER-LB", "Lichtblauwe trigger", "VERPAKKING", "stuks", 1200, 500, 0, "Distrifill"),
                ("ETI-UNI-600", "Etiket Universol 600 ml", "ETIKET", "stuks", 1480, 500, 0, "Leverancier etiketten"),
            ]
            for code, omschrijving, type_, eenheid, voorraad, minimum, kostprijs, leverancier in seed_items:
                upsert_item(db, code=code, omschrijving=omschrijving, type=type_, eenheid=eenheid, voorraad=voorraad, minimum=minimum, kostprijs=kostprijs, leverancier=leverancier)
            for naam in ["Copagro", "TABS Bouwwinkel DC", "Olijslager", "Claasen Coatings", "Verfwinkel.nl", "Volders"]:
                db.add(Customer(naam=naam))
            for naam, type_ in [("Distrifill", "Afvuller"), ("Julius Hoesch", "Grondstoffen"), ("Superdoos", "Dozen"), ("PCC", "Surfactants")]:
                db.add(Supplier(naam=naam, type=type_))
            db.add(Movement(item_code="UNI-CAP-MICHAEL", mutatie=37800, bron="Startdata", referentie="Michael overzicht", gebruiker="Ron"))
            db.commit()
    finally:
        db.close()


def norm(value: Any) -> str:
    return str(value or "").strip().lower().replace("\xa0", " ")


def as_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace("€", "").replace(" ", "")
    if not s:
        return None
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    else:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def detect_code(row_values: list[Any]) -> str | None:
    joined = " | ".join(norm(v) for v in row_values)
    for code, aliases in ALIASES.items():
        if norm(code) in joined:
            return code
        if any(alias in joined for alias in aliases):
            return code
    return None


def likely_numbers(row_values: list[Any]) -> list[float]:
    nums = []
    for v in row_values:
        n = as_float(v)
        if n is None or n in {0, 1}:
            continue
        nums.append(n)
    return nums


def parse_excel(path: Path) -> dict:
    wb = load_workbook(path, data_only=True, read_only=True)
    rows_read = 0
    found = []
    unknown_text_rows = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            rows_read += 1
            values = list(row)
            if not any(v is not None and str(v).strip() for v in values):
                continue
            code = detect_code(values)
            text = " | ".join(str(v).strip() for v in values if v is not None and str(v).strip())
            if code:
                nums = likely_numbers(values)
                quantity = max(nums) if nums else 0
                price_candidates = [n for n in nums if 0 < n < 10000 and n != quantity]
                unit_price = min(price_candidates) if price_candidates else 0
                found.append({
                    "sheet": ws.title,
                    "code": code,
                    "description": text[:300],
                    "quantity": quantity,
                    "unit_price": unit_price,
                    "raw_row": text[:1000],
                })
            elif len(text) > 6 and len(unknown_text_rows) < 20:
                unknown_text_rows.append({"sheet": ws.title, "text": text[:300]})
    return {"rows_read": rows_read, "found": found, "unknown_examples": unknown_text_rows}


app = FastAPI(title="GROBÉ OS", version=APP_VERSION)


@app.on_event("startup")
def on_startup():
    seed()


@app.get("/", response_class=HTMLResponse)
def index():
    return HTML


@app.get("/api/health")
def health():
    return {"status": "ok", "programma": "GROBÉ OS", "versie": APP_VERSION}


@app.get("/api/items")
def items(db: Session = Depends(get_db)):
    return [item_dict(r) for r in db.query(Item).order_by(Item.type, Item.omschrijving).all()]


@app.get("/api/dashboard")
def dashboard(db: Session = Depends(get_db)):
    rows = [item_dict(r) for r in db.query(Item).order_by(Item.type, Item.omschrijving).all()]
    groups = {"finished_stock": [], "tanks": [], "production_capacity": [], "raw_materials": [], "packaging": [], "labels": []}
    for r in rows:
        groups.setdefault(TYPE_GROEP.get(r["type"], "packaging"), []).append(r)
    moves = db.query(Movement).order_by(Movement.datum.desc()).limit(5).all()
    groups.update({
        "finance": {"debtors": "Nog niet gekoppeld", "revenue_month": "Nog niet gekoppeld"},
        "recent_activity": [{"action": m.bron, "description": f"{m.item_code}: {m.mutatie:g} {m.referentie or ''}"} for m in moves],
        "totals": {"items": len(rows), "below_minimum": sum(1 for r in rows if r["status"] == "bijna_op")},
    })
    return groups


class MovementIn(BaseModel):
    item_code: str
    mutatie: float
    bron: str
    referentie: Optional[str] = None
    gebruiker: str = "Ron"


@app.post("/api/movement")
def add_movement(data: MovementIn, db: Session = Depends(get_db)):
    item = db.query(Item).filter(Item.code == data.item_code).first()
    if not item:
        raise HTTPException(status_code=404, detail="Artikel niet gevonden")
    if item.voorraad + data.mutatie < 0:
        raise HTTPException(status_code=400, detail="Onvoldoende voorraad")
    item.voorraad += data.mutatie
    db.add(Movement(**data.model_dump()))
    db.commit()
    return {"success": True, "nieuwe_voorraad": item.voorraad}


@app.get("/api/movements")
def movements(db: Session = Depends(get_db)):
    rows = db.query(Movement).order_by(Movement.datum.desc()).limit(100).all()
    return [{"item_code": r.item_code, "mutatie": r.mutatie, "bron": r.bron, "referentie": r.referentie, "gebruiker": r.gebruiker, "datum": r.datum.isoformat() if r.datum else None} for r in rows]


@app.get("/api/search")
def search(q: str = "", db: Session = Depends(get_db)):
    term = f"%{q}%"
    products = db.query(Item).filter(or_(Item.code.ilike(term), Item.omschrijving.ilike(term), Item.type.ilike(term))).limit(20).all() if q else []
    customers = db.query(Customer).filter(Customer.naam.ilike(term)).limit(20).all() if q else []
    suppliers = db.query(Supplier).filter(Supplier.naam.ilike(term)).limit(20).all() if q else []
    return {"products": [item_dict(p) for p in products], "customers": [{"id": c.id, "name": c.naam, "email": c.email} for c in customers], "suppliers": [{"id": s.id, "name": s.naam, "type": s.type} for s in suppliers]}


@app.get("/api/imports")
def imports(db: Session = Depends(get_db)):
    rows = db.query(ImportLog).order_by(ImportLog.datum.desc()).limit(50).all()
    return [{"source": r.source, "filename": r.filename, "rows_read": r.rows_read, "rows_changed": r.rows_changed, "message": r.message, "datum": r.datum.isoformat() if r.datum else None} for r in rows]


@app.post("/api/import/{source}")
async def import_file(source: str, file: UploadFile = File(...), db: Session = Depends(get_db)):
    suffix = Path(file.filename or "").suffix.lower()
    if suffix not in {".xlsx", ".xlsm"}:
        msg = f"Bestandstype {suffix} wordt nog niet inhoudelijk gelezen."
        db.add(ImportLog(source=source, filename=file.filename or "bestand", rows_read=0, rows_changed=0, message=msg))
        db.commit()
        return {"success": True, "message": msg, "filename": file.filename, "recognized": [], "changed": 0}
    with NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        tmp.write(await file.read())
        tmp_path = Path(tmp.name)
    try:
        parsed = parse_excel(tmp_path)
    finally:
        try:
            tmp_path.unlink()
        except FileNotFoundError:
            pass
    found = parsed.get("found", [])
    changed = 0
    price_changes = []
    for line in found:
        item = db.query(Item).filter(Item.code == line["code"]).first()
        if not item:
            continue
        qty = float(line.get("quantity") or 0)
        price = float(line.get("unit_price") or 0)
        if qty > 0:
            item.voorraad += qty
            db.add(Movement(item_code=item.code, mutatie=qty, bron=f"Import {source}", referentie=file.filename))
            changed += 1
        if price > 0 and abs((item.kostprijs or 0) - price) > 0.0001:
            price_changes.append({"code": item.code, "oud": item.kostprijs or 0, "nieuw": price})
            db.add(PriceHistory(item_code=item.code, supplier=item.leverancier or source, old_price=item.kostprijs or 0, new_price=price, source=source, reference=file.filename))
            item.kostprijs = price
        db.add(PurchaseLine(item_code=item.code, description=line.get("description", ""), quantity=qty, unit_price=price, supplier=item.leverancier or source, filename=file.filename or "", raw_row=line.get("raw_row", "")))
    message = f"Excel gelezen. {len(found)} herkende regels gevonden, {changed} voorraadregels geboekt. Controleer altijd het importresultaat." if found else "Bestand gelezen, maar geen bekende Daniel/Michael regels herkend."
    db.add(ImportLog(source=source, filename=file.filename or "bestand", rows_read=int(parsed.get("rows_read") or 0), rows_changed=changed, message=message))
    db.commit()
    return {"success": True, "message": message, "filename": file.filename, "rows_read": parsed.get("rows_read", 0), "recognized": found[:100], "changed": changed, "price_changes": price_changes, "unknown_examples": parsed.get("unknown_examples", [])[:10]}


@app.get("/api/price-history")
def price_history(db: Session = Depends(get_db)):
    rows = db.query(PriceHistory).order_by(PriceHistory.datum.desc()).limit(100).all()
    return [{"item_code": r.item_code, "supplier": r.supplier, "old_price": r.old_price, "new_price": r.new_price, "source": r.source, "reference": r.reference, "datum": r.datum.isoformat() if r.datum else None} for r in rows]


@app.get("/api/purchase-lines")
def purchase_lines(db: Session = Depends(get_db)):
    rows = db.query(PurchaseLine).order_by(PurchaseLine.datum.desc()).limit(100).all()
    return [{"item_code": r.item_code, "description": r.description, "quantity": r.quantity, "unit_price": r.unit_price, "supplier": r.supplier, "filename": r.filename, "datum": r.datum.isoformat() if r.datum else None} for r in rows]


class FillIn(BaseModel):
    product_code: str = "UNI-5L"
    quantity_units: float = 10
    tank_code: str = "UNI-TANK"


@app.post("/api/fill")
def fill(data: FillIn, db: Session = Depends(get_db)):
    product = db.query(Item).filter(Item.code == data.product_code).first()
    tank = db.query(Item).filter(Item.code == data.tank_code).first()
    if not product or not tank:
        raise HTTPException(status_code=404, detail="Product of tank niet gevonden")
    liters_per_unit = 5 if "5L" in product.code else 1
    liters = data.quantity_units * liters_per_unit
    if tank.voorraad < liters:
        raise HTTPException(status_code=400, detail="Onvoldoende tankvoorraad")
    tank.voorraad -= liters
    product.voorraad += data.quantity_units
    db.add(Movement(item_code=tank.code, mutatie=-liters, bron="Afvulling", referentie=product.code))
    db.add(Movement(item_code=product.code, mutatie=data.quantity_units, bron="Afvulling", referentie=tank.code))
    db.commit()
    return {"success": True, "product_voorraad": product.voorraad, "tank_voorraad": tank.voorraad}


class PackingSlipLine(BaseModel):
    product_code: str
    quantity_units: float


class PackingSlipIn(BaseModel):
    number: str
    customer: str
    lines: list[PackingSlipLine]


@app.post("/api/packing-slips")
def packing_slip(data: PackingSlipIn, db: Session = Depends(get_db)):
    changed = []
    for line in data.lines:
        item = db.query(Item).filter(Item.code == line.product_code).first()
        if not item:
            raise HTTPException(status_code=404, detail=f"Artikel niet gevonden: {line.product_code}")
        if item.voorraad < line.quantity_units:
            raise HTTPException(status_code=400, detail=f"Onvoldoende voorraad: {line.product_code}")
        item.voorraad -= line.quantity_units
        db.add(Movement(item_code=item.code, mutatie=-line.quantity_units, bron="Pakbon", referentie=f"{data.number} - {data.customer}"))
        changed.append({"code": item.code, "nieuwe_voorraad": item.voorraad})
    db.commit()
    return {"success": True, "changed": changed}


HTML = r'''<!doctype html><html lang="nl"><head><meta charset="utf-8"/><meta name="viewport" content="width=device-width,initial-scale=1"/><title>GROBÉ OS</title><style>
:root{--dark:#2E3C7A;--light:#8FBCE7;--orange:#F5A623;--red:#D32F2F;--bg:#F7F9FC;--card:#fff;--line:#E4EAF2;--text:#2F3A45;--muted:#6B7280}*{box-sizing:border-box}body{margin:0;font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",sans-serif;background:var(--bg);color:var(--text)}header{height:72px;background:#fff;border-bottom:1px solid var(--line);display:flex;align-items:center;justify-content:space-between;padding:0 32px;position:sticky;top:0;z-index:2}h1{font-size:24px;margin:0;color:var(--dark)}nav{display:flex;gap:8px;flex-wrap:wrap}.navbtn{background:#fff;color:var(--dark);border:1px solid var(--line)}.pill{border:1px solid var(--line);border-radius:999px;padding:8px 14px;color:var(--muted);background:#fff}.wrap{padding:28px;display:grid;grid-template-columns:1fr 1.5fr 1fr;gap:22px}.card{background:var(--card);border:1px solid var(--line);border-radius:18px;padding:20px;box-shadow:0 6px 18px rgba(46,60,122,.05);margin-bottom:18px}.card h2{font-size:18px;margin:0 0 16px;color:var(--dark)}.row{display:flex;align-items:center;justify-content:space-between;padding:12px 0;border-bottom:1px solid #EEF2F7}.row:last-child{border-bottom:0}.name{font-weight:650}.sub{font-size:13px;color:var(--muted);margin-top:3px}.dot{width:15px;height:15px;border-radius:50%;display:inline-block;margin-right:10px;vertical-align:-2px;border:1px solid rgba(0,0,0,.08)}button{background:var(--dark);color:#fff;border:0;border-radius:12px;padding:10px 14px;font-weight:650;cursor:pointer}input,select{width:100%;padding:10px;border:1px solid var(--line);border-radius:10px;margin:6px 0 12px;background:#fff}.small{font-size:13px;color:var(--muted)}.hidden{display:none}.grid2{display:grid;grid-template-columns:1fr 1fr;gap:12px}.full{grid-column:1/-1}pre{white-space:pre-wrap;background:#F2F5FA;padding:12px;border-radius:12px;overflow:auto}@media(max-width:1000px){.wrap{grid-template-columns:1fr}header{height:auto;align-items:flex-start;gap:12px;flex-direction:column;padding:16px}}
</style></head><body><header><div><h1>GROBÉ OS</h1><div class="small">v0.8 live start · echte basis, geen eindversie</div></div><nav><button class="navbtn" onclick="show('dashboard')">Dashboard</button><button class="navbtn" onclick="show('import')">Import</button><button class="navbtn" onclick="show('prices')">Prijzen</button><button class="navbtn" onclick="show('search')">Zoeken</button><button class="navbtn" onclick="show('log')">Logboek</button></nav></header>
<main id="dashboard" class="wrap page"><section><div class="card"><h2>Vandaag</h2><div id="finance" class="small"></div></div><div class="card"><h2>Laatste activiteit</h2><div id="activity"></div></div></section><section><div class="card"><h2>Gereed product</h2><div id="finished"></div></div><div class="card"><h2>Tanks</h2><div id="tanks"></div></div><div class="card"><h2>Productiecapaciteit Michael</h2><div id="capacity"></div></div><div class="card"><h2>Grondstoffen</h2><div id="raw"></div></div></section><section><div class="card"><h2>Verpakkingen</h2><div id="packaging"></div></div><div class="card"><h2>Etiketten</h2><div id="labels"></div></div><div class="card"><h2>Acties</h2><div class="grid2"><button onclick="demoFill()">Afvullen 10x 5L</button><button onclick="demoPakbon()">Pakbon 4x 5L</button></div><p class="small">Schrijft echte voorraadmutaties en auditlog.</p></div></section></main>
<main id="import" class="wrap page hidden"><section class="card full"><h2>Importcentrum</h2><p class="small">Eerste parser. Controle blijft nodig. PDF-inhoud wordt nog niet gelezen.</p><div class="grid2"><div><h3>Daniel Excel</h3><input type="file" id="danielFile"><button onclick="upload('daniel')">Importeer Daniel</button></div><div><h3>Michael Excel</h3><input type="file" id="michaelFile"><button onclick="upload('michael')">Importeer Michael</button></div></div><pre id="importResult" class="small"></pre></section><section class="card full"><h2>Importlog</h2><div id="imports"></div></section></main>
<main id="prices" class="wrap page hidden"><section class="card full"><h2>Prijshistorie</h2><div id="priceHistory"></div></section><section class="card full"><h2>Inkoopregels</h2><div id="purchaseLines"></div></section></main>
<main id="search" class="wrap page hidden"><section class="card full"><h2>Centrale zoekfunctie</h2><input id="q" placeholder="Zoek product, klant of leverancier" oninput="doSearch()"><div id="searchResults"></div></section></main>
<main id="log" class="wrap page hidden"><section class="card full"><h2>Voorraadmutaties</h2><div id="movements"></div></section></main>
<script>
async function api(path, opts){const r=await fetch(path, opts); if(!r.ok){alert(await r.text()); throw new Error(r.status)} return r.json()}
function row(name, sub, status){return `<div class="row"><div><div class="name"><span class="dot" style="background:${status?.color||'#8FBCE7'}"></span>${name}</div><div class="sub">${sub}${status?.label?' · '+status.label:''}</div></div></div>`}
function simple(items, f){return items.map(f).join('') || '<div class="small">Geen regels.</div>'}
async function load(){const d=await api('/api/dashboard');finance.innerHTML=`Open debiteuren: <b>${d.finance.debtors}</b><br>Omzet maand: <b>${d.finance.revenue_month}</b><br>Artikelen: <b>${d.totals.items}</b>`;finished.innerHTML=simple(d.finished_stock,x=>row(x.omschrijving,`${x.voorraad} ${x.eenheid} · min ${x.minimum}`,{color:x.kleur,label:x.status}));tanks.innerHTML=simple(d.tanks,x=>row(x.omschrijving,`${x.voorraad} liter`,{color:x.kleur,label:x.status}));capacity.innerHTML=simple(d.production_capacity,x=>row(x.omschrijving,`${x.voorraad} liter · bron: ${x.leverancier}`,{color:x.kleur,label:x.status}));raw.innerHTML=simple(d.raw_materials,x=>row(x.omschrijving,`${x.voorraad} ${x.eenheid} · € ${Number(x.kostprijs||0).toFixed(2)}`,{color:x.kleur,label:x.status}));packaging.innerHTML=simple(d.packaging,x=>row(x.omschrijving,`${x.voorraad} ${x.eenheid}`,{color:x.kleur,label:x.status}));labels.innerHTML=simple(d.labels,x=>row(x.omschrijving,`${x.voorraad} ${x.eenheid}`,{color:x.kleur,label:x.status}));activity.innerHTML=simple(d.recent_activity,x=>`<div class="row"><div><div class="name">${x.action}</div><div class="sub">${x.description}</div></div></div>`);loadImports();loadLogs();loadPrices()}
async function loadImports(){const d=await api('/api/imports');imports.innerHTML=simple(d,x=>`<div class="row"><div><div class="name">${x.source} · ${x.filename}</div><div class="sub">gelezen ${x.rows_read}, gewijzigd ${x.rows_changed} · ${x.message}</div></div></div>`)}
async function loadLogs(){const m=await api('/api/movements');movements.innerHTML=simple(m,x=>`<div class="row"><div><div class="name">${x.bron} ${x.mutatie}</div><div class="sub">${x.item_code} · ${x.referentie||''}</div></div></div>`)}
async function loadPrices(){const h=await api('/api/price-history');priceHistory.innerHTML=simple(h,x=>`<div class="row"><div><div class="name">${x.item_code}</div><div class="sub">${x.supplier} · € ${x.old_price} → € ${x.new_price} · ${x.reference||''}</div></div></div>`);const p=await api('/api/purchase-lines');purchaseLines.innerHTML=simple(p,x=>`<div class="row"><div><div class="name">${x.item_code} · ${x.quantity}</div><div class="sub">€ ${x.unit_price} · ${x.filename}<br>${x.description}</div></div></div>`)}
function show(id){document.querySelectorAll('.page').forEach(p=>p.classList.add('hidden'));document.getElementById(id).classList.remove('hidden');if(id==='log')loadLogs();if(id==='import')loadImports();if(id==='prices')loadPrices()}
async function demoFill(){await api('/api/fill',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({product_code:'UNI-5L',quantity_units:10,tank_code:'UNI-TANK'})});load()}
async function demoPakbon(){await api('/api/packing-slips',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({number:'TEST-'+Date.now(),customer:'Test klant',lines:[{product_code:'UNI-5L',quantity_units:4}]})});load()}
async function upload(kind){const inp=document.getElementById(kind+'File');if(!inp.files[0]){alert('Kies eerst een bestand');return}const fd=new FormData();fd.append('file',inp.files[0]);const r=await api('/api/import/'+kind,{method:'POST',body:fd});importResult.textContent=JSON.stringify(r,null,2);load()}
let timer;function doSearch(){clearTimeout(timer);timer=setTimeout(async()=>{if(q.value.length<2){searchResults.innerHTML='';return}const d=await api('/api/search?q='+encodeURIComponent(q.value));searchResults.innerHTML='<h3>Producten</h3>'+simple(d.products,x=>`<div class="row"><div><div class="name">${x.omschrijving}</div><div class="sub">${x.code} · ${x.type}</div></div></div>`)+'<h3>Klanten</h3>'+simple(d.customers,x=>`<div class="row"><div><div class="name">${x.name}</div></div></div>`)+'<h3>Leveranciers</h3>'+simple(d.suppliers,x=>`<div class="row"><div><div class="name">${x.name}</div><div class="sub">${x.type}</div></div></div>`);},250)}
load();
</script></body></html>'''
